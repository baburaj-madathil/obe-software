"""
CO-WK / CO-PO / CO-PSO Mapping utility for NBA / Washington Accord OBE analysis.

Two pipelines:
  1. generate_co_wk_excel   – PDF syllabus → Course Outcomes + WK mapping
  2. generate_co_po_mapping  – Excel of COs → CO-PO/PSO strength matrix + justifications

Requires: google-genai, pypdf, pandas, openpyxl
"""

from __future__ import annotations

import os
import re
import json
import logging
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

try:
    from google import genai
    from google.genai import types
except ImportError:
    genai = None
    types = None

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Reference data (Washington Accord Knowledge Profiles WK1–WK9)
# ---------------------------------------------------------------------------

WK_FULL_DATA = [
    {
        "WK": "WK1",
        "Title": "Natural Sciences & Social Sciences",
        "Aspects": "Physics, Chemistry, Biology, Geology, Social Science awareness.",
        "Focus": "Scientific explanation of materials, ecosystems, environment.",
    },
    {
        "WK": "WK2",
        "Title": "Math, Data & Computing",
        "Aspects": "Calculus, Statistics, Numerical analysis, Algorithmic thinking, Programming (Python/MATLAB), Data modelling.",
        "Focus": "Data interpretation, modelling, regression, or trend analysis.",
    },
    {
        "WK": "WK3",
        "Title": "Engineering Fundamentals",
        "Aspects": "Mechanics, Thermodynamics, Fluid mechanics, formulating problems using basic laws.",
        "Focus": "Fundamental laws governing engineering systems.",
    },
    {
        "WK": "WK4",
        "Title": "Discipline-Specific Knowledge",
        "Aspects": "Professional codes, standards, advanced tools, specialist theoretical frameworks.",
        "Focus": "Application of professional standards and codes of practice.",
    },
    {
        "WK": "WK5",
        "Title": "Sustainability & Economics",
        "Aspects": "Environmental impacts, carbon footprint, life-cycle cost, resource reuse, net zero.",
        "Focus": "Sustainable design, green technologies or environmental protection.",
    },
    {
        "WK": "WK6",
        "Title": "Engineering Practice & Technology",
        "Aspects": "Industry tools, materials, lab/fieldwork, practical procedures, hands-on skills.",
        "Focus": "Understanding of processes, materials, equipment, or technologies.",
    },
    {
        "WK": "WK7",
        "Title": "Engineer's Role in Society",
        "Aspects": "Public safety, legal responsibility, societal impact, community needs.",
        "Focus": "Infrastructure for society, disaster resilience.",
    },
    {
        "WK": "WK8",
        "Title": "Research & Innovation",
        "Aspects": "Literature review, critical thinking, creative approaches, emerging issues.",
        "Focus": "Exposure to latest research or trends.",
    },
    {
        "WK": "WK9",
        "Title": "Ethics & Inclusivity",
        "Aspects": "Professional ethics, integrity, diversity, inclusive behavior.",
        "Focus": "Professional conduct and ethical decision-making.",
    },
]

ANALYSIS_SCHEMA = {
    "type": "ARRAY",
    "items": {
        "type": "OBJECT",
        "properties": {
            "CO_Number": {"type": "STRING"},
            "Course_Outcome": {"type": "STRING"},
            "Blooms_Level": {"type": "STRING"},
            "Mapped_WKs": {"type": "ARRAY", "items": {"type": "STRING"}},
            "Primary_WK": {"type": "STRING"},
            "Detailed_Justification": {
                "type": "STRING",
                "description": "Comprehensive explanation of why these WKs were mapped based on the CO text and WK Key Aspects.",
            },
        },
        "required": ["CO_Number", "Course_Outcome", "Blooms_Level", "Mapped_WKs", "Primary_WK", "Detailed_Justification"],
    },
}

DEFAULT_MODEL = "gemini-2.5-flash"
DEFAULT_INPUT_EXCEL = "CO_WK_Mapping.xlsx"
DEFAULT_OUTPUT_EXCEL = "CO_PO_Mapping.xlsx"

# ---------------------------------------------------------------------------
# GAPC V4.0 Base Program Outcomes (PO1 - PO11)
# ---------------------------------------------------------------------------

BASE_PO_CATALOG: dict[str, dict] = {
    "PO1": {
        "title": "Engineering Knowledge",
        "statement": "Apply knowledge of mathematics, natural science, computing, engineering fundamentals and an engineering specialization to develop solutions for complex engineering problems.",
        "wks": ["WK1", "WK2", "WK3", "WK4"],
        "pis": {
            "1.1.1": "Apply mathematical techniques such as calculus, linear algebra, and statistics to solve problems",
            "1.1.2": "Apply advanced mathematical techniques to model and solve engineering problems",
            "1.2.1": "Apply laws of natural science to an engineering problem",
            "1.3.1": "Apply fundamental engineering concepts to solve engineering problems",
            "1.4.1": "Apply specialized engineering knowledge of the program to solve engineering problems",
        },
    },
    "PO2": {
        "title": "Problem Analysis",
        "statement": "Identify, formulate, review research literature and analyze complex engineering problems reaching substantiated conclusions.",
        "wks": ["WK1", "WK2", "WK3", "WK4"],
        "pis": {
            "2.1.1": "Articulate problem statements and identify objectives",
            "2.1.2": "Identify engineering systems, variables, and parameters to solve the problems",
            "2.1.3": "Identify the mathematical, engineering and relevant knowledge that applies to a given problem",
            "2.2.1": "Reframe complex problems into interconnected sub-problems",
            "2.2.2": "Identify, assemble and evaluate information and resources",
            "2.2.3": "Identify existing processes/solution methods including justified approximations and assumptions",
            "2.2.4": "Compare alternative solution processes balancing technical, socio-economic and environmental dimensions",
            "2.3.1": "Combine scientific principles and engineering concepts to formulate model/s of a system or process",
            "2.3.2": "Identify necessary assumptions to allow modeling of a system at required accuracy",
            "2.4.1": "Apply engineering mathematics and computations to solve mathematical models",
            "2.4.2": "Produce and validate results through skillful use of contemporary engineering tools and models",
            "2.4.3": "Identify sources of error in the solution process and limitations of the solution",
            "2.4.4": "Extract desired understanding and conclusions consistent with objectives and limitations",
        },
    },
    "PO3": {
        "title": "Design/Development of Solutions",
        "statement": "Design creative solutions for complex engineering problems and design systems/components/processes to meet identified needs.",
        "wks": ["WK5"],
        "pis": {
            "3.1.1": "Recognize that need analysis is key to good problem definition",
            "3.1.2": "Elicit and document engineering requirements from stakeholders",
            "3.1.3": "Synthesize engineering requirements from a review of the state-of-the-art",
            "3.1.4": "Extract engineering requirements from relevant Codes and Standards",
            "3.1.5": "Explore requirements considering health, safety, environmental, cultural and societal issues",
            "3.1.6": "Determine design objectives, functional requirements and specifications",
            "3.2.1": "Apply formal idea generation tools to develop multiple engineering design solutions",
            "3.2.2": "Build models/prototypes to develop a diverse set of design solutions",
            "3.2.3": "Identify suitable criteria for the evaluation of alternate design solutions",
            "3.3.1": "Apply decision-making tools to select optimal design solutions",
            "3.3.2": "Consult with domain experts and stakeholders to select candidate engineering design solutions",
            "3.4.1": "Refine conceptual design into detailed design considering public safety, whole-life cost, and environment",
        },
    },
    "PO4": {
        "title": "Conduct Investigations of Complex Problems",
        "statement": "Conduct investigations of complex engineering problems using research-based knowledge and design of experiments.",
        "wks": ["WK8"],
        "pis": {
            "4.1.1": "Define a problem, its scope and importance for investigation",
            "4.1.2": "Examine relevant methods, tools and techniques of experiment design, data acquisition and presentation",
            "4.1.3": "Apply appropriate instrumentation and software tools to make measurements",
            "4.1.4": "Establish a relationship between measured data and underlying physical principles",
            "4.2.1": "Design and develop an experimental approach, specify equipment and procedures",
            "4.2.2": "Understand the statistical design of experiments and choose an appropriate plan",
            "4.3.1": "Use appropriate procedures, tools and techniques to conduct experiments and collect data",
            "4.3.2": "Analyze data for trends and correlations, stating possible errors and limitations",
            "4.3.3": "Represent data to facilitate analysis, explanation and drawing of conclusions",
        },
    },
    "PO5": {
        "title": "Engineering Tool Usage",
        "statement": "Create, select and apply appropriate techniques, resources and modern engineering and IT tools.",
        "wks": ["WK2", "WK6"],
        "pis": {
            "5.1.1": "Identify modern engineering tools such as computer-aided drafting, modeling and analysis tools",
            "5.1.2": "Create/adapt/modify/extend tools and techniques to solve engineering problems",
            "5.2.1": "Identify strengths and limitations of tools for modeling, simulating, and monitoring system performance",
            "5.2.2": "Demonstrate proficiency in using discipline-specific tools",
            "5.3.1": "Discuss limitations and validate tools, techniques and resources",
            "5.3.2": "Verify credibility of results with reference to accuracy, limitations, and assumptions",
        },
    },
    "PO6": {
        "title": "The Engineer and The World",
        "statement": "Analyze and evaluate societal and environmental aspects while solving complex engineering problems.",
        "wks": ["WK1", "WK5", "WK7"],
        "pis": {
            "6.1.1": "Identify engineering roles for impact on sustainability, health, safety, legal framework, and culture",
            "6.2.1": "Interpret legislation, regulations, codes, and standards relevant to the discipline",
            "6.3.1": "Identify risks/impacts in the life-cycle of an engineering product or activity",
            "6.3.2": "Understand relationship between technical, socio-economic and environmental dimensions of sustainability",
        },
    },
    "PO7": {
        "title": "Ethics",
        "statement": "Apply ethical principles and commit to professional ethics, human values, diversity and inclusion.",
        "wks": ["WK9"],
        "pis": {
            "7.1.1": "Identify situations of unethical professional conduct and propose ethical alternatives",
            "7.1.2": "Understand the need for diversity with mutual respect and inclusive attitudes",
            "7.2.1": "Identify tenets of the professional code of ethics",
            "7.2.2": "Examine and apply moral and ethical principles to known case studies",
        },
    },
    "PO8": {
        "title": "Individual and Collaborative Team Work",
        "statement": "Function effectively as an individual, and as a member or leader in diverse/multi-disciplinary teams.",
        "wks": [],
        "pis": {
            "8.1.1": "Recognize a variety of working preferences and appreciate team diversity",
            "8.1.2": "Implement norms of practice for effective teamwork to accomplish goals",
            "8.2.1": "Demonstrate communication, problem-solving, conflict resolution and leadership skills",
            "8.2.2": "Treat team members respectfully",
            "8.2.3": "Listen to other team members",
            "8.2.4": "Maintain composure in difficult situations",
            "8.3.1": "Present results as a team with smooth integration of individual efforts",
        },
    },
    "PO9": {
        "title": "Communication",
        "statement": "Communicate effectively and inclusively within the engineering community and society at large.",
        "wks": [],
        "pis": {
            "9.1.1": "Read, understand and interpret technical and non-technical information",
            "9.1.2": "Produce clear, well-constructed, and well-supported written engineering documents",
            "9.1.3": "Create a logical progression of ideas in reports and presentations",
            "9.2.1": "Listen to and comprehend information, instructions, and viewpoints of others",
            "9.2.2": "Deliver effective oral presentations to technical and non-technical audiences",
            "9.3.1": "Create engineering-standard figures, reports and drawings to complement writing",
            "9.3.2": "Use a variety of media effectively to convey messages in documents or presentations",
        },
    },
    "PO10": {
        "title": "Project Management and Finance",
        "statement": "Apply engineering management principles and economic decision-making to manage projects.",
        "wks": [],
        "pis": {
            "10.1.1": "Describe economic and financial costs/benefits of an engineering activity",
            "10.1.2": "Analyze financial statements to evaluate project financial status",
            "10.2.1": "Analyze and select appropriate proposals based on economic and financial considerations",
            "10.3.1": "Identify tasks and resources required to complete an engineering activity",
            "10.3.2": "Use project management tools to schedule projects within budget and time constraints",
        },
    },
    "PO11": {
        "title": "Life Long Learning",
        "statement": "Recognize the need for and have the ability for independent and life-long learning.",
        "wks": ["WK8"],
        "pis": {
            "11.1.1": "Describe the rationale for continuing professional development",
            "11.1.2": "Identify deficiencies in knowledge and source information to close the gap",
            "11.2.1": "Identify technological advances in engineering requiring continuous education",
            "11.2.2": "Recognize the need to stay current regarding new developments in your field",
            "11.3.1": "Source and comprehend technical literature and credible sources of information",
            "11.3.2": "Analyze sourced technical information for feasibility, viability, and sustainability",
        },
    },
}

# ---------------------------------------------------------------------------
# Load PSOs and associated WKs from PSOs.txt
# ---------------------------------------------------------------------------

def load_psos_from_file(pso_file_path: str = "PSOs.txt") -> dict[str, dict]:
    psos = {}
    path = Path(pso_file_path)
    
    if path.exists():
        print(f"📄 Found custom PSOs file: '{pso_file_path}'. Parsing Program Specific Outcomes & WKs...")
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if ":" in line:
                    pso_id, text = line.split(":", 1)
                    pso_id = pso_id.strip().upper()
                    text = text.strip()
                    
                    wks = ["WK3", "WK4", "WK6"]
                    wk_match = re.search(r"\[(.*?)\]", text)
                    if wk_match:
                        raw_wks = wk_match.group(1)
                        parsed_wks = [w.strip().upper() for w in raw_wks.split(",") if w.strip()]
                        if parsed_wks:
                            wks = parsed_wks
                        text = re.sub(r"\[.*?\]", "", text).strip()
                    
                    statement = text
                    psos[pso_id] = {
                        "title": f"{pso_id} (Program Specific Outcome)",
                        "statement": statement,
                        "wks": wks,
                        "pis": {
                            f"{pso_id}.1": f"Demonstrate mastery of core concepts ({', '.join(wks)}) for {statement.lower()}",
                            f"{pso_id}.2": f"Apply specialized domain tools and methods for {statement.lower()}",
                        },
                    }
    else:
        psos = {
            "PSO1": {
                "title": "Program Specific Outcome 1 (Core Engineering & System Design)",
                "statement": "Apply domain-specific engineering principles to analyze, design, and implement technical systems.",
                "wks": ["WK3", "WK4"],
                "pis": {
                    "PSO1.1": "Apply domain knowledge to solve core technical and engineering design challenges",
                    "PSO1.2": "Design and develop domain-specific hardware, software, or process systems",
                },
            },
            "PSO2": {
                "title": "Program Specific Outcome 2 (Modern Tools & Professional Practice)",
                "statement": "Utilize modern industry-standard software tools, frameworks, and experimental methods.",
                "wks": ["WK2", "WK6"],
                "pis": {
                    "PSO2.1": "Apply specialized software tools, frameworks, and platforms for problem solving",
                    "PSO2.2": "Demonstrate proficiency in practical execution and technical experimentation",
                },
            },
            "PSO3": {
                "title": "Program Specific Outcome 3 (Societal, Environment & Industrial Solutions)",
                "statement": "Address multidisciplinary, environmental, and societal engineering challenges.",
                "wks": ["WK5", "WK7"],
                "pis": {
                    "PSO3.1": "Incorporate sustainability, environmental impact, or industrial standards in engineering solutions",
                    "PSO3.2": "Formulate solutions for real-world multidisciplinary and societal problems",
                },
            },
        }
    return psos


def get_full_catalog() -> dict[str, dict]:
    catalog = dict(BASE_PO_CATALOG)
    custom_psos = load_psos_from_file()
    catalog.update(custom_psos)
    return catalog


def gapc_mapping_value(yes_count: int, total_pis: int) -> tuple[int, float]:
    if total_pis <= 0 or yes_count <= 0:
        return 0, 0.0
    x = (yes_count / total_pis) * 100.0
    if x < 10:
        return 0, round(x, 1)
    if x <= 33:
        return 1, round(x, 1)
    if x <= 67:
        return 2, round(x, 1)
    return 3, round(x, 1)


# ---------------------------------------------------------------------------
# Styling & Helpers
# ---------------------------------------------------------------------------

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
s3_fill = PatternFill("solid", fgColor="63BE7B")
s2_fill = PatternFill("solid", fgColor="FFEB84")
s1_fill = PatternFill("solid", fgColor="F4B183")
no_fill = PatternFill("solid", fgColor="D9D9D9")
error_fill = PatternFill("solid", fgColor="FF6B6B")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def _resolve_api_key(api_key: str) -> str:
    if not api_key or not str(api_key).strip() or str(api_key).strip() in ("", "YOUR_GEMINI_API_KEY_HERE"):
        raise ValueError("api_key is required. Pass a valid Gemini API key.")
    return str(api_key).strip()


def _strip_markdown_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
    if text.endswith("```"):
        text = text[:-3].strip()
    return text


def _po_sort_key(po_name: str) -> tuple[int, int]:
    if po_name.startswith("PSO"):
        num = po_name[3:]
        return (2, int(num) if num.isdigit() else 99)
    elif po_name.startswith("PO"):
        num = po_name[2:]
        return (1, int(num) if num.isdigit() else 99)
    return (3, 99)


def extract_syllabus_text(pdf_path: str) -> str:
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"Source PDF file not found at: {pdf_path}")
    reader = PdfReader(pdf_path)
    return "\n".join([p.extract_text() for p in reader.pages if p.extract_text()])


# ---------------------------------------------------------------------------
# Pipeline 1: PDF → CO + WK mapping
# ---------------------------------------------------------------------------

def generate_co_wk_excel(
    pdf_path: str,
    model_name: str,
    api_key: str,
    output_excel_path: str = DEFAULT_INPUT_EXCEL,
) -> pd.DataFrame:
    if not model_name or not str(model_name).strip():
        raise ValueError("model_name is required.")
    if genai is None:
        raise ImportError("Required dependency 'google-genai' is missing.")

    effective_api_key = _resolve_api_key(api_key)
    syllabus_text = extract_syllabus_text(pdf_path)
    client = genai.Client(api_key=effective_api_key)

    prompt = (
        "You are an expert Accreditation Auditor. Perform a dual task:\n"
        "1. Generate 5 precise, measurable Course Outcomes (CO1-CO5) based on the syllabus provided.\n"
        "2. For each CO, map EVERY possible Washington Accord Knowledge Profile (WK1-WK9) that applies.\n"
        "3. Provide a 'Detailed_Justification' for each CO.\n\n"
        f"WK DEFINITIONS:\n{json.dumps(WK_FULL_DATA, indent=2)}\n\n"
        f"SYLLABUS CONTENT:\n{syllabus_text}"
    )

    response = client.models.generate_content(
        model=model_name,
        contents=prompt,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=ANALYSIS_SCHEMA,
        ),
    )

    raw_data = response.parsed
    table_rows = []
    for item in raw_data:
        get = item.get if isinstance(item, dict) else lambda k, default="": getattr(item, k, default)
        mapped_wks = get("Mapped_WKs", []) or []
        if not isinstance(mapped_wks, (list, tuple)): mapped_wks = [str(mapped_wks)]

        table_rows.append({
            "CO ID": get("CO_Number", ""),
            "Course Outcome": get("Course_Outcome", ""),
            "Bloom's Level": get("Blooms_Level", ""),
            "Mapped WKs": ", ".join(str(w) for w in mapped_wks),
            "Primary WK": get("Primary_WK", ""),
            "Justification": get("Detailed_Justification", ""),
        })

    df = pd.DataFrame(table_rows)
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="CO_WK_Mapping", index=False)

    return df


# ---------------------------------------------------------------------------
# Pipeline 2: CO-PO / PSO mapping using Gemini
# ---------------------------------------------------------------------------

def init_gemini_client(api_key: str):
    if genai is None: raise ImportError("Required dependency 'google-genai' is missing.")
    return genai.Client(api_key=_resolve_api_key(api_key))


def analyze_co_with_gemini(
    client,
    co_id: str,
    statement: str,
    bloom: str,
    existing_wks: str,
    model_name: str,
    catalog: dict[str, dict] = None,
) -> dict:

    if catalog is None:
        catalog = get_full_catalog()

    pi_catalog_text = []
    for po, info in catalog.items():
        wks_str = f" [Associated WKs: {', '.join(info['wks'])}]" if info.get("wks") else ""
        pi_lines = [f"  {pid}: {desc}" for pid, desc in info["pis"].items()]
        pi_catalog_text.append(f"{po} ({info['title']}){wks_str} – {len(info['pis'])} PIs:\n" + "\n".join(pi_lines))
    catalog_block = "\n\n".join(pi_catalog_text)

    all_target_names = ", ".join(catalog.keys())

    system_prompt = f"""
You are an expert NBA & Washington Accord Accreditation Auditor.

Your MANDATE is to produce an EXHAUSTIVE evaluation of ONE Course Outcome (CO)
against EVERY SINGLE Outcome: {all_target_names}.

MAPPING GUIDELINES (MAXIMIZE COVERAGE):
1. SYSTEMATICALLY CHECK ALL OUTCOMES ({all_target_names}). Do NOT restrict checks only to primary outcomes.
2. Pay close attention to associated Knowledge Profiles (WKs) listed for each PO/PSO. If a CO matches a WK associated with a PO or PSO, thoroughly evaluate if that outcome's PIs are supported.
3. Mark a Performance Indicator (PI) as "Yes" whenever the CO statement explicitly OR implicitly supports that skill or action.

Return ONLY a valid JSON object:
{{
  "co_id": "{co_id}",
  "understanding": "Detailed breakdown of student expectations",
  "suggested_wks": ["WK1", "WK2", "WK3"],
  "primary_wk": "WK3",
  "pi_yes": {{
    "PO1": ["1.1.1"],
    "PSO1": ["PSO1.1"]
  }},
  "justifications": {{
    "PO1": "Supports PO1 because...",
    "PSO1": "Supports PSO1 because..."
  }},
  "recommendations": "Pedagogical recommendations"
}}

PERFORMANCE INDICATOR CATALOG:
{catalog_block}
"""

    prompt = f"{system_prompt}\n\nCO ID: {co_id}\nStatement: {statement}\nBloom's Level: {bloom}\nExisting WK mapping: {existing_wks}\n\nReturn ONLY valid JSON."

    try:
        response = client.models.generate_content(
            model=model_name,
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.3,
                response_mime_type="application/json",
            ),
        )
        text = _strip_markdown_fences(response.text or "")
        raw = json.loads(text)

        pi_yes = raw.get("pi_yes") or {}
        justifications = raw.get("justifications") or {}
        mapped_pos: dict[str, dict] = {}

        for po, yes_list in pi_yes.items():
            if po not in catalog: continue
            catalog_pis = catalog[po]["pis"]
            valid_yes = [pid for pid in (yes_list or []) if pid in catalog_pis]
            total = len(catalog_pis)
            yes_count = len(valid_yes)
            strength, x_pct = gapc_mapping_value(yes_count, total)
            if strength == 0: continue
            mapped_pos[po] = {
                "strength": strength,
                "mapping_strength_pct": x_pct,
                "yes_count": yes_count,
                "total_pis": total,
                "contributing_pis": valid_yes,
                "justification": justifications.get(po, ""),
            }

        return {
            "co_id": raw.get("co_id", co_id),
            "understanding": raw.get("understanding", ""),
            "suggested_wks": raw.get("suggested_wks") or [],
            "primary_wk": raw.get("primary_wk", ""),
            "mapped_pos": mapped_pos,
            "recommendations": raw.get("recommendations", ""),
        }
    except Exception as e:
        logger.exception("Error analyzing %s", co_id)
        return {"co_id": co_id, "error": str(e)}


def read_cos_from_excel(excel_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    df = pd.read_excel(excel_path, sheet_name=sheet_name) if sheet_name else pd.read_excel(excel_path)
    col_map = {}
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if "co id" in col_lower or col_lower == "co": col_map[col] = "CO ID"
        elif "course outcome" in col_lower or "statement" in col_lower: col_map[col] = "Course Outcome"
        elif "bloom" in col_lower: col_map[col] = "Bloom's Level"
        elif "mapped wk" in col_lower or col_lower in ("mapped_wks", "wks"): col_map[col] = "Mapped WKs"
    return df.rename(columns=col_map)


def analyze_all_cos(excel_path: str, client, model_name: str, sheet_name: Optional[str] = None, catalog: dict[str, dict] = None) -> list[dict]:
    if catalog is None: catalog = get_full_catalog()
    df = read_cos_from_excel(excel_path, sheet_name)
    results = []

    print(f"\nAnalyzing {len(df)} COs against {len(catalog)} POs/PSOs using Gemini ({model_name})...\n")

    for _, row in df.iterrows():
        co_id = str(row["CO ID"]).strip()
        statement = str(row["Course Outcome"]).strip()
        bloom = str(row.get("Bloom's Level", "") or "").strip()
        existing_wks = str(row.get("Mapped WKs", "") or "").strip()

        print(f"  Analyzing {co_id} ...")
        res = analyze_co_with_gemini(client, co_id, statement, bloom, existing_wks, model_name=model_name, catalog=catalog)
        results.append(res)
        print(f"  → {'Done' if 'error' not in res else 'ERROR'}")

    return results


def _compute_strength_stats(results: list[dict], all_pos: list[str]) -> dict:
    co_stats = []
    po_stats = {po: {"count_3": 0, "count_2": 0, "count_1": 0, "total_mapped": 0, "sum_strength": 0} for po in all_pos}
    total_mappings = sum_all_strengths = errors = 0

    for res in results:
        co = res.get("co_id", "")
        if "error" in res:
            errors += 1
            co_stats.append({"co_id": co, "mapped_count": 0, "sum_strength": 0, "avg_strength": None, "count_3": 0, "count_2": 0, "count_1": 0, "error": True})
            continue

        mapped = res.get("mapped_pos") or {}
        c3 = c2 = c1 = ssum = 0
        for po in all_pos:
            data = mapped.get(po) or {}
            strength = data.get("strength")
            if strength not in (1, 2, 3): continue
            ssum += strength; total_mappings += 1; sum_all_strengths += strength
            po_stats[po]["total_mapped"] += 1; po_stats[po]["sum_strength"] += strength
            if strength == 3: c3 += 1; po_stats[po]["count_3"] += 1
            elif strength == 2: c2 += 1; po_stats[po]["count_2"] += 1
            else: c1 += 1; po_stats[po]["count_1"] += 1

        mapped_count = c3 + c2 + c1
        co_stats.append({"co_id": co, "mapped_count": mapped_count, "sum_strength": ssum, "avg_strength": round(ssum / mapped_count, 2) if mapped_count else None, "count_3": c3, "count_2": c2, "count_1": c1, "error": False})

    for po, st in po_stats.items():
        st["avg_strength"] = round(st["sum_strength"] / st["total_mapped"], 2) if st["total_mapped"] else None

    return {
        "co_stats": co_stats, "po_stats": po_stats,
        "overall": {
            "total_cos": len(results), "errors": errors, "total_mappings": total_mappings,
            "sum_all_strengths": sum_all_strengths,
            "avg_strength": round(sum_all_strengths / total_mappings, 2) if total_mappings else None,
            "pos_covered": sum(1 for st in po_stats.values() if st["total_mapped"] > 0),
        }
    }


def create_excel_report(results: list[dict], model_name: str, output_path: str = DEFAULT_OUTPUT_EXCEL, input_excel: Optional[str] = None, catalog: dict[str, dict] = None) -> None:
    if catalog is None: catalog = get_full_catalog()
    wb = Workbook()
    ALL_POS = sorted(list(catalog.keys()), key=_po_sort_key)
    stats = _compute_strength_stats(results, ALL_POS)

    # ----- Sheet 0: Source CO-WK Data -----
    if input_excel and os.path.exists(input_excel):
        try:
            src_df = pd.read_excel(input_excel)
            ws0 = wb.active
            ws0.title = "0_Source_CO_WK"
            ws0["A1"] = f"Source CO-WK Mapping (from {os.path.basename(input_excel)})"
            ws0["A1"].font = Font(bold=True, size=13, color="1F4E79")
            ws0.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(src_df.columns), 1))

            for col_idx, col_name in enumerate(src_df.columns, 1):
                cell = ws0.cell(row=3, column=col_idx, value=str(col_name))
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

            for r_idx, row in enumerate(src_df.itertuples(index=False), 4):
                for c_idx, value in enumerate(row, 1):
                    cell = ws0.cell(row=r_idx, column=c_idx, value=value if pd.notna(value) else "")
                    cell.border = thin_border; cell.alignment = left_align if c_idx > 1 else center
                ws0.row_dimensions[r_idx].height = 40

            ws0.column_dimensions["A"].width = 12
            ws0.column_dimensions["B"].width = 55
            ws0.column_dimensions["C"].width = 15
            ws0.column_dimensions["D"].width = 22
            ws0.column_dimensions["E"].width = 15
            ws0.column_dimensions["F"].width = 55

            ws1 = wb.create_sheet("1_Strength_Matrix")
        except Exception:
            ws1 = wb.active; ws1.title = "1_Strength_Matrix"
    else:
        ws1 = wb.active; ws1.title = "1_Strength_Matrix"

    # ----- Sheet 1: Strength Matrix -----
    ws1["A1"] = "CO-PO/PSO Correlation Strength Matrix (GAPC V4.0)"
    ws1["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ALL_POS) + 1)

    ws1["A2"] = "3 = Substantial (X=68–100%) | 2 = Moderate (X=34–67%) | 1 = Slight (X=10–33%) | – = No correlation"
    ws1["A2"].font = Font(italic=True, size=9)

    ws1.cell(row=4, column=1, value="CO")
    for j, po in enumerate(ALL_POS, 2): ws1.cell(row=4, column=j, value=po)
    _style_header(ws1, 4, 1, len(ALL_POS) + 1)

    for i, res in enumerate(results):
        r = 5 + i
        co = res.get("co_id", f"CO{i + 1}")
        cell_co = ws1.cell(row=r, column=1, value=co)
        cell_co.font = Font(bold=True); cell_co.alignment = center; cell_co.border = thin_border
        mapped = res.get("mapped_pos") or {}
        for j, po in enumerate(ALL_POS, 2):
            cell = ws1.cell(row=r, column=j)
            strength = (mapped.get(po) or {}).get("strength")
            if strength == 3: cell.value = 3; cell.fill = s3_fill
            elif strength == 2: cell.value = 2; cell.fill = s2_fill
            elif strength == 1: cell.value = 1; cell.fill = s1_fill
            else: cell.value = "–"; cell.fill = no_fill
            cell.alignment = center; cell.border = thin_border; cell.font = Font(bold=True)

    ws1.column_dimensions["A"].width = 10
    for col in range(2, len(ALL_POS) + 2):
        ws1.column_dimensions[get_column_letter(col)].width = 8

    # ----- Sheet 2: Strength Calculation -----
    ws_calc = wb.create_sheet("2_Strength_Calculation")
    ws_calc["A1"] = "CO-PO/PSO Mapping Strength Calculation (GAPC V4.0)"
    ws_calc["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws_calc.merge_cells("A1:H1")

    ws_calc["A4"] = "A. Per-CO Strength Calculation"
    ws_calc["A4"].font = Font(bold=True, size=11, color="1F4E79")

    co_headers = ["CO", "Outcomes Mapped", "Count (3)", "Count (2)", "Count (1)", "Sum of Strengths", "Average Strength", "Status"]
    for col, h in enumerate(co_headers, 1): ws_calc.cell(row=5, column=col, value=h)
    _style_header(ws_calc, 5, 1, 8)

    for i, cs in enumerate(stats["co_stats"]):
        r = 6 + i
        ws_calc.cell(row=r, column=1, value=cs["co_id"]).alignment = center
        ws_calc.cell(row=r, column=2, value=cs["mapped_count"]).alignment = center
        ws_calc.cell(row=r, column=3, value=cs["count_3"]).alignment = center
        ws_calc.cell(row=r, column=4, value=cs["count_2"]).alignment = center
        ws_calc.cell(row=r, column=5, value=cs["count_1"]).alignment = center
        ws_calc.cell(row=r, column=6, value=cs["sum_strength"]).alignment = center
        avg_cell = ws_calc.cell(row=r, column=7, value=cs["avg_strength"] if cs["avg_strength"] is not None else "–")
        avg_cell.alignment = center
        status_cell = ws_calc.cell(row=r, column=8, value="OK" if not cs.get("error") else "ERROR")
        status_cell.alignment = center
        for c in range(1, 9): ws_calc.cell(row=r, column=c).border = thin_border

    po_start = 6 + len(stats["co_stats"]) + 2
    ws_calc.cell(row=po_start, column=1, value="B. Per-PO/PSO Strength Calculation").font = Font(bold=True, size=11, color="1F4E79")
    po_headers = ["PO / PSO", "COs Mapped", "Count (3)", "Count (2)", "Count (1)", "Sum of Strengths", "Average Strength", "Coverage"]
    header_row = po_start + 1
    for col, h in enumerate(po_headers, 1): ws_calc.cell(row=header_row, column=col, value=h)
    _style_header(ws_calc, header_row, 1, 8)

    for i, po in enumerate(ALL_POS):
        r = header_row + 1 + i
        st = stats["po_stats"][po]
        ws_calc.cell(row=r, column=1, value=po).alignment = center
        ws_calc.cell(row=r, column=2, value=st["total_mapped"]).alignment = center
        ws_calc.cell(row=r, column=3, value=st["count_3"]).alignment = center
        ws_calc.cell(row=r, column=4, value=st["count_2"]).alignment = center
        ws_calc.cell(row=r, column=5, value=st["count_1"]).alignment = center
        ws_calc.cell(row=r, column=6, value=st["sum_strength"]).alignment = center
        avg_cell = ws_calc.cell(row=r, column=7, value=st["avg_strength"] if st["avg_strength"] is not None else "–")
        avg_cell.alignment = center
        cov_cell = ws_calc.cell(row=r, column=8, value="Yes" if st["total_mapped"] > 0 else "No")
        cov_cell.alignment = center
        for c in range(1, 9): ws_calc.cell(row=r, column=c).border = thin_border

    overall_start = header_row + 1 + len(ALL_POS) + 2
    ws_calc.cell(row=overall_start, column=1, value="C. Overall Strength Metrics").font = Font(bold=True, size=11, color="1F4E79")
    ov = stats["overall"]
    metrics = [
        ("Total COs analyzed", ov["total_cos"]),
        ("COs with errors", ov["errors"]),
        ("Total CO-PO/PSO mappings", ov["total_mappings"]),
        ("Sum of all strengths", ov["sum_all_strengths"]),
        ("Overall average strength", ov["avg_strength"] if ov["avg_strength"] is not None else "–"),
        ("POs/PSOs covered", f"{ov['pos_covered']} / {len(ALL_POS)}"),
    ]
    for i, (label, value) in enumerate(metrics):
        r = overall_start + 1 + i
        ws_calc.cell(row=r, column=1, value=label).border = thin_border
        val_cell = ws_calc.cell(row=r, column=2, value=value)
        val_cell.alignment = center; val_cell.border = thin_border

    ws_calc.column_dimensions["A"].width = 22
    ws_calc.column_dimensions["B"].width = 16
    ws_calc.column_dimensions["C"].width = 12
    ws_calc.column_dimensions["D"].width = 12
    ws_calc.column_dimensions["E"].width = 12
    ws_calc.column_dimensions["F"].width = 16
    ws_calc.column_dimensions["G"].width = 16
    ws_calc.column_dimensions["H"].width = 12

    # ----- Sheet 3: Justifications -----
    ws2 = wb.create_sheet("3_Mapping_with_Justification")
    ws2["A1"] = "Complete CO-PO/PSO Mapping with Justification"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2.merge_cells("A1:H1")

    headers2 = ["CO", "PO / PSO", "Strength", "X (%)", "Yes/Total PIs", "Contributing PIs", "Justification", "Primary WK"]
    for col, h in enumerate(headers2, 1): ws2.cell(row=3, column=col, value=h)
    _style_header(ws2, 3, 1, 8)

    row_num = 4
    for res in results:
        co = res.get("co_id", "")
        primary_wk = res.get("primary_wk", "")
        mapped = res.get("mapped_pos") or {}
        for po, data in sorted(mapped.items(), key=lambda x: _po_sort_key(x[0])):
            ws2.cell(row=row_num, column=1, value=co).alignment = center
            ws2.cell(row=row_num, column=2, value=po).alignment = center
            ws2.cell(row=row_num, column=3, value=data.get("strength")).alignment = center
            ws2.cell(row=row_num, column=4, value=data.get("mapping_strength_pct")).alignment = center
            ws2.cell(row=row_num, column=5, value=f"{data.get('yes_count')}/{data.get('total_pis')}").alignment = center
            ws2.cell(row=row_num, column=6, value=", ".join(data.get("contributing_pis") or [])).alignment = center
            ws2.cell(row=row_num, column=7, value=data.get("justification", "")).alignment = left_align
            ws2.cell(row=row_num, column=8, value=primary_wk).alignment = center
            for c in range(1, 9): ws2.cell(row=row_num, column=c).border = thin_border
            ws2.row_dimensions[row_num].height = 50
            row_num += 1

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 28
    ws2.column_dimensions["G"].width = 70
    ws2.column_dimensions["H"].width = 12

    # ----- Sheet 4: Understanding & Recommendations -----
    ws3 = wb.create_sheet("4_Understanding_Recommendations")
    ws3["A1"] = "Deep Understanding of each CO + Recommendations"
    ws3["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws3.merge_cells("A1:D1")

    headers3 = ["CO", "Understanding of CO", "Suggested WKs", "Recommendations"]
    for col, h in enumerate(headers3, 1): ws3.cell(row=3, column=col, value=h)
    _style_header(ws3, 3, 1, 4)

    for i, res in enumerate(results):
        r = 4 + i
        ws3.cell(row=r, column=1, value=res.get("co_id", "")).alignment = center
        ws3.cell(row=r, column=2, value=res.get("understanding", "")).alignment = left_align
        suggested = res.get("suggested_wks") or []
        ws3.cell(row=r, column=3, value=", ".join(str(w) for w in suggested)).alignment = center
        ws3.cell(row=r, column=4, value=res.get("recommendations", "")).alignment = left_align
        for c in range(1, 5): ws3.cell(row=r, column=c).border = thin_border
        ws3.row_dimensions[r].height = 70

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 65
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 50

    # ----- Sheet 5: PI Yes/No Detail -----
    ws_pi = wb.create_sheet("5_PI_YesNo_Detail")
    ws_pi["A1"] = "PO/PSO–Competency–PI–CO Mapping Status (GAPC V4.0 Step 6)"
    ws_pi["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws_pi.merge_cells("A1:G1")

    co_ids = [res.get("co_id", f"CO{i+1}") for i, res in enumerate(results)]
    yes_lookup = {res.get("co_id", ""): {po: set(data.get("contributing_pis") or []) for po, data in (res.get("mapped_pos") or {}).items()} for res in results}

    row = 4
    for po in ALL_POS:
        info = catalog.get(po)
        if not info: continue
        ws_pi.cell(row=row, column=1, value=f"{po}: {info['title']}").font = Font(bold=True, size=11, color="1F4E79")
        ws_pi.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + len(co_ids))
        row += 1

        headers = ["Competency", "PI ID", "Performance Indicator"] + co_ids
        for col, h in enumerate(headers, 1): ws_pi.cell(row=row, column=col, value=h)
        _style_header(ws_pi, row, 1, len(headers))
        row += 1

        for pid, pdesc in info["pis"].items():
            parts = pid.split(".")
            competency = ".".join(parts[:2]) if len(parts) >= 2 else pid
            ws_pi.cell(row=row, column=1, value=competency).alignment = center
            ws_pi.cell(row=row, column=2, value=pid).alignment = center
            ws_pi.cell(row=row, column=3, value=pdesc).alignment = left_align
            for j, cid in enumerate(co_ids):
                is_yes = pid in yes_lookup.get(cid, {}).get(po, set())
                cell = ws_pi.cell(row=row, column=4 + j, value="Yes" if is_yes else "No")
                cell.alignment = center; cell.fill = s3_fill if is_yes else no_fill
            for c in range(1, 4 + len(co_ids)): ws_pi.cell(row=row, column=c).border = thin_border
            row += 1

        ws_pi.cell(row=row, column=1, value="Mapping Strength X (%)").font = Font(bold=True)
        ws_pi.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        for j, cid in enumerate(co_ids):
            yes_c = len(yes_lookup.get(cid, {}).get(po, set()))
            _, x_pct = gapc_mapping_value(yes_c, len(info["pis"]))
            cell = ws_pi.cell(row=row, column=4 + j, value=x_pct)
            cell.alignment = center; cell.font = Font(bold=True); cell.border = thin_border
        row += 1

        ws_pi.cell(row=row, column=1, value="Mapping Value (1/2/3)").font = Font(bold=True)
        ws_pi.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        for j, cid in enumerate(co_ids):
            yes_c = len(yes_lookup.get(cid, {}).get(po, set()))
            strength, _ = gapc_mapping_value(yes_c, len(info["pis"]))
            cell = ws_pi.cell(row=row, column=4 + j, value=strength if strength else "–")
            cell.alignment = center; cell.font = Font(bold=True); cell.border = thin_border
            if strength == 3: cell.fill = s3_fill
            elif strength == 2: cell.fill = s2_fill
            elif strength == 1: cell.fill = s1_fill
            else: cell.fill = no_fill
        row += 2

    ws_pi.column_dimensions["A"].width = 14
    ws_pi.column_dimensions["B"].width = 10
    ws_pi.column_dimensions["C"].width = 70
    for j in range(len(co_ids)): ws_pi.column_dimensions[get_column_letter(4 + j)].width = 8

    # ----- Sheet 6: Summary -----
    ws4 = wb.create_sheet("6_Summary")
    ws4["A1"] = "Summary"
    ws4["A1"].font = Font(bold=True, size=13, color="1F4E79")

    covered = [po for po, st in stats["po_stats"].items() if st["total_mapped"] > 0]
    ws4["A3"] = "POs/PSOs covered:"; ws4["B3"] = ", ".join(sorted(covered, key=_po_sort_key)) if covered else "None"
    ws4["A4"] = "Model used:"; ws4["B4"] = model_name
    ws4["A5"] = "Total COs analyzed:"; ws4["B5"] = ov["total_cos"]
    ws4["A6"] = "COs with errors:"; ws4["B6"] = ov["errors"]
    ws4["A7"] = "Total CO-PO/PSO mappings:"; ws4["B7"] = ov["total_mappings"]
    ws4["A8"] = "Overall average strength:"; ws4["B8"] = ov["avg_strength"] if ov["avg_strength"] is not None else "–"

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 70

    wb.save(output_path)
    print(f"\nExcel report saved → {output_path}")


def generate_co_po_mapping(model_name: str, api_key: str, input_excel: str = DEFAULT_INPUT_EXCEL, output_excel: str = DEFAULT_OUTPUT_EXCEL, sheet_name: Optional[str] = None) -> list[dict]:
    catalog = get_full_catalog()
    client = init_gemini_client(api_key)
    results = analyze_all_cos(input_excel, client, model_name=model_name, sheet_name=sheet_name, catalog=catalog)
    
    json_path = "gemini_raw_analysis.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    create_excel_report(results, model_name=model_name, output_path=output_excel, input_excel=input_excel, catalog=catalog)
    return results


def main_cli():
    import argparse
    parser = argparse.ArgumentParser(description="CO-WK / CO-PO / CO-PSO mapping for OBE")
    sub = parser.add_subparsers(dest="command")

    p1 = sub.add_parser("wk")
    p1.add_argument("pdf"); p1.add_argument("--model", required=True); p1.add_argument("--api-key", required=True); p1.add_argument("-o", "--output", default=DEFAULT_INPUT_EXCEL)

    p2 = sub.add_parser("po")
    p2.add_argument("excel", nargs="?", default=DEFAULT_INPUT_EXCEL); p2.add_argument("--model", required=True); p2.add_argument("--api-key", required=True); p2.add_argument("-o", "--output", default=DEFAULT_OUTPUT_EXCEL); p2.add_argument("--sheet", default=None)

    args = parser.parse_args()
    if args.command == "wk": generate_co_wk_excel(args.pdf, args.model, args.api_key, args.output)
    elif args.command == "po": generate_co_po_mapping(args.model, args.api_key, args.excel, args.output, args.sheet)
    else: parser.print_help()


if __name__ == "__main__":
    main_cli()
